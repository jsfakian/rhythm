#!/usr/bin/env python3
"""
RHYTHM server-assigned Repository Study ID manifest creator WITH DICOM UID extraction.

The partner metadata template contains filenames and manual metadata only. This tool
opens each ZIP file, extracts the anonymized DICOM StudyInstanceUID (0020,000D), and
includes it in the JSON manifest as `dicom_uid`. The server still assigns the final
RHYTHM Repository Study ID after validating the manifest and ZIP contents.

Template columns:
  filename, site_code, clinical_indication_code, anatomical_region, contrast_code,
  patient_group_code, scanner_id, protocol_name, patient_weight_kg,
  patient_age_years, ctdivol_mgy, dlp_mgy_cm, image_quality

Prerequisites:
  pip install pydicom openpyxl

This tool does not anonymize DICOM files. Use it only with already anonymized CT ZIPs.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import queue
import sys
import threading
import traceback
import zipfile
from pathlib import Path
from typing import Any, Dict, List, Tuple

TOOL_VERSION = "3.1"

REQUIRED_COLUMNS = [
    "filename",
    "site_code",
    "clinical_indication_code",
    "anatomical_region",
    "contrast_code",
    "patient_group_code",
    "scanner_id",
    "protocol_name",
    "patient_weight_kg",
    "patient_age_years",
    "ctdivol_mgy",
    "dlp_mgy_cm",
    "image_quality",
]

FORBIDDEN_COLUMNS = [
    "repo_id",
    "repository_study_id",
    "reserved_repo_id",
    "patient_name",
    "name",
    "surname",
    "original_patient_id",
    "hospital_patient_id",
    "accession_number",
    "birth_date",
    "date_of_birth",
    "national_id",
    "social_security_number",
]


def try_openpyxl():
    try:
        import openpyxl  # type: ignore
        return openpyxl
    except ImportError:
        return None


def require_pydicom():
    try:
        import pydicom  # type: ignore
        return pydicom
    except ImportError as exc:
        raise RuntimeError(
            "pydicom is required to extract DICOM UIDs from ZIP files. "
            "Install it with: pip install pydicom"
        ) from exc


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def to_number_or_none(value: Any) -> Any:
    text = clean(value)
    if not text:
        return None
    text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return text


def file_sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for block in iter(lambda: f.read(1024 * 1024), b""):
            h.update(block)
    return h.hexdigest()


def is_zip_file(path: Path) -> bool:
    return path.exists() and path.is_file() and zipfile.is_zipfile(path)


def inspect_dicom_zip(zip_path: Path) -> Dict[str, Any]:
    """Extract DICOM UID summary from one ZIP containing one anonymized CT studyset."""
    pydicom = require_pydicom()
    patient_ids = set()
    study_uids = set()
    series_uids = set()
    sop_uids = set()
    readable_dicom_files = 0
    skipped_files = 0

    with zipfile.ZipFile(zip_path, "r") as z:
        for info in z.infolist():
            if info.is_dir():
                continue
            lower_name = info.filename.lower()
            if lower_name.endswith((".txt", ".json", ".csv", ".xml", ".jpg", ".jpeg", ".png", ".bmp", ".pdf")):
                skipped_files += 1
                continue
            try:
                with z.open(info, "r") as fp:
                    ds = pydicom.dcmread(fp, stop_before_pixels=True, force=False)
            except Exception:
                skipped_files += 1
                continue

            study_uid = clean(getattr(ds, "StudyInstanceUID", ""))
            if not study_uid:
                skipped_files += 1
                continue

            readable_dicom_files += 1
            study_uids.add(study_uid)

            patient_id = clean(getattr(ds, "PatientID", ""))
            series_uid = clean(getattr(ds, "SeriesInstanceUID", ""))
            sop_uid = clean(getattr(ds, "SOPInstanceUID", ""))

            if patient_id:
                patient_ids.add(patient_id)
            if series_uid:
                series_uids.add(series_uid)
            if sop_uid:
                sop_uids.add(sop_uid)

    if readable_dicom_files == 0:
        raise ValueError(f"No readable DICOM files with StudyInstanceUID found in ZIP: {zip_path.name}")
    if len(study_uids) > 1:
        raise ValueError(
            f"ZIP contains multiple StudyInstanceUID values and should be split into one ZIP per studyset: "
            f"{zip_path.name} -> {sorted(study_uids)}"
        )
    if len(patient_ids) > 1:
        raise ValueError(
            f"ZIP contains multiple DICOM PatientID values under one study and should be reviewed: "
            f"{zip_path.name} -> {sorted(patient_ids)}"
        )

    study_uid = sorted(study_uids)[0]
    patient_id = sorted(patient_ids)[0] if patient_ids else ""
    return {
        "patient_id": patient_id,
        "study_uid": study_uid,
        "uid_type": "StudyInstanceUID",
        "series_uids": sorted(series_uids),
        "series_count": len(series_uids),
        "instance_count": len(sop_uids),
        "readable_dicom_files": readable_dicom_files,
        "skipped_files": skipped_files,
    }


def load_csv(path: Path) -> Tuple[List[Dict[str, str]], List[str]]:
    with path.open("r", newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        headers = [clean(h) for h in (reader.fieldnames or [])]
        rows = [{clean(k): clean(v) for k, v in row.items()} for row in reader]
    return rows, headers


def load_xlsx(path: Path) -> Tuple[List[Dict[str, str]], List[str]]:
    openpyxl = try_openpyxl()
    if openpyxl is None:
        raise RuntimeError("openpyxl is required for .xlsx input. Install with: pip install openpyxl")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    raw_rows = list(ws.iter_rows(values_only=True))
    if not raw_rows:
        raise ValueError("Excel file is empty.")
    headers = [clean(h) for h in raw_rows[0]]
    rows = []
    for raw in raw_rows[1:]:
        if not any(clean(x) for x in raw):
            continue
        row = {}
        for i, header in enumerate(headers):
            if header:
                row[header] = clean(raw[i] if i < len(raw) else "")
        rows.append(row)
    return rows, headers


def load_metadata(path: Path) -> Tuple[List[Dict[str, str]], List[str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return load_csv(path)
    if suffix in [".xlsx", ".xlsm"]:
        return load_xlsx(path)
    raise ValueError("Unsupported metadata file type. Use .csv or .xlsx.")


def check_rows(rows: List[Dict[str, str]], headers: List[str], zip_folder: Path) -> List[str]:
    errors: List[str] = []
    missing_cols = [c for c in REQUIRED_COLUMNS if c not in headers]
    if missing_cols:
        errors.append("Missing required columns: " + ", ".join(missing_cols))
    forbidden_present = [c for c in FORBIDDEN_COLUMNS if c in headers]
    if forbidden_present:
        errors.append("Remove these columns from the server-assigned template: " + ", ".join(forbidden_present))
    if not rows:
        errors.append("No metadata rows found.")

    seen_filenames = set()
    for idx, row in enumerate(rows, start=2):
        for col in REQUIRED_COLUMNS:
            if col in headers and not clean(row.get(col, "")):
                errors.append(f"Row {idx}: missing required value '{col}'")
        filename = clean(row.get("filename", ""))
        if filename:
            if filename in seen_filenames:
                errors.append(f"Row {idx}: duplicate filename in template: {filename}")
            seen_filenames.add(filename)
            path = zip_folder / filename
            if not path.exists():
                errors.append(f"Row {idx}: ZIP file not found in selected ZIP folder: {filename}")
            elif not path.is_file():
                errors.append(f"Row {idx}: filename is not a file: {filename}")
            elif path.suffix.lower() != ".zip":
                errors.append(f"Row {idx}: file is not named as a .zip file: {filename}")
            elif not is_zip_file(path):
                errors.append(f"Row {idx}: file is not a valid ZIP archive: {filename}")
    return errors


def build_manifest(
    metadata_file: Path,
    zip_folder: Path,
    output_json: Path,
    batch_id: str = "",
    include_sha256: bool = True,
) -> Dict[str, Any]:
    rows, headers = load_metadata(metadata_file)
    if not zip_folder.exists() or not zip_folder.is_dir():
        raise ValueError(f"ZIP folder does not exist or is not a directory: {zip_folder}")

    errors = check_rows(rows, headers, zip_folder)
    if errors:
        raise ValueError("\n".join(errors[:100]) + ("\n..." if len(errors) > 100 else ""))

    site_codes = sorted(set(clean(r.get("site_code", "")) for r in rows))
    site = site_codes[0] if len(site_codes) == 1 else "MULTISITE"
    batch = clean(batch_id) or f"{site}-BATCH001"

    items = []
    index_rows = []
    seen_study_uids = set()

    for i, row in enumerate(rows, start=1):
        filename = clean(row.get("filename"))
        zip_path = zip_folder / filename
        dicom_summary = inspect_dicom_zip(zip_path)
        dicom_uid = dicom_summary["study_uid"]
        if dicom_uid in seen_study_uids:
            raise ValueError(
                f"Duplicate StudyInstanceUID found across different ZIP files: {dicom_uid}. "
                "Each study should be submitted only once."
            )
        seen_study_uids.add(dicom_uid)

        item = {
            "ref": f"ROW{i:04d}",
            "filename": filename,
            "dicom_uid": dicom_uid,
            "dicom": dicom_summary,
            "site_code": clean(row.get("site_code")),
            "clinical_indication_code": clean(row.get("clinical_indication_code")),
            "anatomical_region": clean(row.get("anatomical_region")),
            "contrast_code": clean(row.get("contrast_code")),
            "patient_group_code": clean(row.get("patient_group_code")),
            "scanner_id": clean(row.get("scanner_id")),
            "protocol_name": clean(row.get("protocol_name")),
            "patient_weight_kg": to_number_or_none(row.get("patient_weight_kg")),
            "patient_age_years": to_number_or_none(row.get("patient_age_years")),
            "ctdivol_mgy": to_number_or_none(row.get("ctdivol_mgy")),
            "dlp_mgy_cm": to_number_or_none(row.get("dlp_mgy_cm")),
            "image_quality": clean(row.get("image_quality")),
            "size_bytes": zip_path.stat().st_size,
        }
        if include_sha256:
            item["sha256"] = file_sha256(zip_path)
        items.append(item)

        index_rows.append({
            "ref": item["ref"],
            "filename": filename,
            "dicom_uid": dicom_uid,
            "dicom_patient_id": dicom_summary.get("patient_id", ""),
            "site_code": item["site_code"],
            "clinical_indication_code": item["clinical_indication_code"],
            "contrast_code": item["contrast_code"],
            "patient_group_code": item["patient_group_code"],
            "size_bytes": item["size_bytes"],
            "sha256": item.get("sha256", ""),
            "series_count": dicom_summary.get("series_count", ""),
            "instance_count": dicom_summary.get("instance_count", ""),
        })

    manifest = {
        "v": "1.0",
        "type": "rhythm_server_assigned_upload_manifest",
        "server_assigns_repo_id": True,
        "site": site,
        "batch": batch,
        "tool": "create_rhythm_server_assigned_manifest_gui_with_uid",
        "tool_version": TOOL_VERSION,
        "note": (
            "Each filename refers to one ZIP file containing one already anonymized CT DICOM studyset. "
            "The partner does not assign repo_id. The server assigns the RHYTHM Repository Study ID after "
            "validating the manifest, ZIP checksum, and DICOM identifiers inside the ZIP. "
            "dicom_uid corresponds to DICOM StudyInstanceUID (0020,000D) extracted from the ZIP."
        ),
        "items": items,
    }

    output_json.parent.mkdir(parents=True, exist_ok=True)
    output_json.write_text(json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")

    index_csv = output_json.with_name(output_json.stem + "_index.csv")
    with index_csv.open("w", newline="", encoding="utf-8-sig") as f:
        fieldnames = [
            "ref", "filename", "dicom_uid", "dicom_patient_id", "site_code",
            "clinical_indication_code", "contrast_code", "patient_group_code",
            "size_bytes", "sha256", "series_count", "instance_count",
        ]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(index_rows)

    return {"manifest": str(output_json), "index_csv": str(index_csv), "items": len(items), "site": site, "batch": batch}


def cli_main() -> None:
    ap = argparse.ArgumentParser(description="Create RHYTHM upload manifest with DICOM StudyInstanceUID. Server assigns repo IDs.")
    ap.add_argument("--input", required=True, help="Metadata CSV/XLSX template.")
    ap.add_argument("--zip-folder", required=True, help="Folder containing ZIP files listed in the filename column.")
    ap.add_argument("--out", required=True, help="Output JSON manifest.")
    ap.add_argument("--batch", default="", help="Optional batch ID.")
    ap.add_argument("--no-sha256", action="store_true", help="Do not compute ZIP SHA-256 checksums.")
    args = ap.parse_args()
    result = build_manifest(Path(args.input), Path(args.zip_folder), Path(args.out), args.batch, not args.no_sha256)
    print(f"Created manifest: {result['manifest']}")
    print(f"Created index CSV: {result['index_csv']}")
    print(f"Items: {result['items']}")
    print("Repository Study IDs will be assigned by the server.")


def run_gui() -> None:
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox

    class App(tk.Tk):
        def __init__(self):
            super().__init__()
            self.title("RHYTHM server-assigned upload manifest creator with DICOM UID")
            self.geometry("960x660")
            self.minsize(860, 560)
            self.q: queue.Queue = queue.Queue()
            self.metadata_var = tk.StringVar()
            self.zip_folder_var = tk.StringVar()
            self.out_var = tk.StringVar()
            self.batch_var = tk.StringVar()
            self.sha_var = tk.BooleanVar(value=True)
            self._build()
            self.after(100, self._poll)

        def _build(self):
            ttk.Label(self, text="RHYTHM upload manifest creator", font=("Segoe UI", 15, "bold")).pack(anchor="w", padx=14, pady=(14, 4))
            ttk.Label(
                self,
                text=("Simplified workflow: metadata template uses filename only. The script extracts DICOM StudyInstanceUID "
                      "from each ZIP and includes it in the manifest. The server assigns the Repository Study ID after upload validation."),
                foreground="#7a3d00",
                wraplength=900,
            ).pack(anchor="w", padx=14, pady=(0, 10))
            frm = ttk.Frame(self)
            frm.pack(fill="x", padx=10)
            self._file_row(frm, 0, "Metadata CSV/XLSX", self.metadata_var, self.browse_metadata)
            self._file_row(frm, 1, "Folder containing ZIP files", self.zip_folder_var, self.browse_zip_folder)
            self._file_row(frm, 2, "Output JSON manifest", self.out_var, self.browse_output)
            pad = {"padx": 10, "pady": 6}
            ttk.Label(frm, text="Batch ID").grid(row=3, column=0, sticky="w", **pad)
            ttk.Entry(frm, textvariable=self.batch_var).grid(row=3, column=1, sticky="ew", **pad)
            ttk.Label(frm, text="Example: S001-2026-07-12-001").grid(row=3, column=2, sticky="w", **pad)
            ttk.Checkbutton(frm, text="Include ZIP SHA-256 checksums in manifest", variable=self.sha_var).grid(row=4, column=1, sticky="w", **pad)
            frm.columnconfigure(1, weight=1)
            btns = ttk.Frame(self)
            btns.pack(fill="x", padx=14, pady=10)
            ttk.Button(btns, text="Generate JSON Manifest", command=self.generate).pack(side="left", padx=(0, 8))
            ttk.Button(btns, text="Clear Log", command=lambda: self.log.delete("1.0", "end")).pack(side="left")
            ttk.Label(self, text="Prerequisites: pip install pydicom openpyxl. This client-side tool does not anonymize DICOM files.", foreground="#475467", wraplength=900).pack(anchor="w", padx=14, pady=(0, 8))
            self.log = tk.Text(self, height=22, wrap="word", font=("Consolas", 10))
            self.log.pack(fill="both", expand=True, padx=14, pady=(0, 14))

        def _file_row(self, parent, row, label, var, command):
            pad = {"padx": 10, "pady": 6}
            ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", **pad)
            ttk.Entry(parent, textvariable=var).grid(row=row, column=1, sticky="ew", **pad)
            ttk.Button(parent, text="Browse...", command=command).grid(row=row, column=2, sticky="e", **pad)

        def browse_metadata(self):
            path = filedialog.askopenfilename(title="Select metadata template", filetypes=[("CSV or Excel", "*.csv *.xlsx *.xlsm"), ("CSV", "*.csv"), ("Excel", "*.xlsx *.xlsm"), ("All files", "*.*")])
            if path:
                self.metadata_var.set(path)
                if not self.out_var.get():
                    p = Path(path)
                    self.out_var.set(str(p.with_name("rhythm_upload_manifest.json")))

        def browse_zip_folder(self):
            path = filedialog.askdirectory(title="Select folder containing ZIP files")
            if path:
                self.zip_folder_var.set(path)

        def browse_output(self):
            path = filedialog.asksaveasfilename(title="Save JSON manifest", defaultextension=".json", initialfile="rhythm_upload_manifest.json", filetypes=[("JSON", "*.json")])
            if path:
                self.out_var.set(path)

        def log_msg(self, msg: str):
            self.log.insert("end", msg + "\n")
            self.log.see("end")
            self.update_idletasks()

        def generate(self):
            if not self.metadata_var.get().strip():
                messagebox.showwarning("Missing metadata", "Select the metadata CSV/XLSX file.")
                return
            if not self.zip_folder_var.get().strip():
                messagebox.showwarning("Missing ZIP folder", "Select the folder containing the ZIP files.")
                return
            if not self.out_var.get().strip():
                messagebox.showwarning("Missing output", "Select the output JSON manifest path.")
                return
            self.log_msg("Creating server-assigned upload manifest with DICOM UID extraction...")
            threading.Thread(target=self._worker, daemon=True).start()

        def _worker(self):
            try:
                result = build_manifest(Path(self.metadata_var.get().strip()), Path(self.zip_folder_var.get().strip()), Path(self.out_var.get().strip()), self.batch_var.get().strip(), self.sha_var.get())
                self.q.put(("success", result))
            except Exception:
                self.q.put(("error", traceback.format_exc()))

        def _poll(self):
            try:
                while True:
                    kind, payload = self.q.get_nowait()
                    if kind == "success":
                        self.log_msg(f"Created manifest: {payload['manifest']}")
                        self.log_msg(f"Created index CSV: {payload['index_csv']}")
                        self.log_msg(f"Items: {payload['items']}")
                        self.log_msg("Repository Study IDs will be assigned by the server.")
                        messagebox.showinfo("Done", "JSON manifest created successfully.")
                    else:
                        self.log_msg("ERROR:")
                        self.log_msg(payload)
                        messagebox.showerror("Error", payload.splitlines()[-1] if payload else "Unknown error")
            except queue.Empty:
                pass
            self.after(100, self._poll)

    App().mainloop()


if __name__ == "__main__":
    if len(sys.argv) > 1 and "--gui" not in sys.argv:
        cli_main()
    else:
        run_gui()
